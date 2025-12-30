import io
import datetime as dt
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string

def parse_legacy_outputs_xlsm(
    xlsm_bytes: bytes,
    sheet_name: str = "Fond_rates_full",
    header_row: int = 3,
    data_start_row: int = 6,
    date_col: str = "Q",
    first_series_col: str = "R",
    last_series_col: str = "AE",
) -> pd.DataFrame:
    """
    Returns LONG dataframe:
      rate_date | series_code | value
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsm_bytes), data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    c_date = column_index_from_string(date_col)
    c1 = column_index_from_string(first_series_col)
    c2 = column_index_from_string(last_series_col)

    series_codes = []
    for c in range(c1, c2 + 1):
        code = ws.cell(row=header_row, column=c).value
        if not code:
            raise ValueError(f"Missing series code at header_row={header_row}, col={c}")
        series_codes.append(str(code).strip())

    rows = []
    r = data_start_row
    while True:
        d = ws.cell(row=r, column=c_date).value
        if d is None:
            break
        if isinstance(d, dt.datetime):
            rate_date = d.date()
        elif isinstance(d, dt.date):
            rate_date = d
        else:
            # skip weird rows
            r += 1
            continue

        for i, code in enumerate(series_codes):
            v = ws.cell(row=r, column=c1 + i).value
            if v is None:
                continue
            rows.append({"rate_date": rate_date, "series_code": code, "value": float(v)})

        r += 1

    return pd.DataFrame(rows)
